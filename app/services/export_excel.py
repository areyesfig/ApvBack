"""Export de simulación APV a Excel."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from app.services.calculadora_apv import simular_regimenes
from app.services.proyeccion import proyectar_jubilacion_anual, proyectar_jubilacion


def _fmt_num(value: float | int | None) -> str:
    if value is None:
        return ""
    return f"{value:,.0f}".replace(",", " ")


def generar_excel_simulacion(
    sueldo_bruto_mensual: float,
    edad_actual: int,
    edad_jubilacion: int,
    ahorro_mensual_apv: float,
    perfil_riesgo: float,
    ahorro_mensual_normal: float,
) -> bytes:
    """Genera un archivo Excel con datos de entrada, totales por régimen y proyección año a año."""
    resultado = simular_regimenes(sueldo_bruto_mensual, ahorro_mensual_apv)
    proyeccion_anual = proyectar_jubilacion_anual(
        ahorro_mensual_apv=ahorro_mensual_apv,
        ahorro_mensual_normal=ahorro_mensual_normal,
        edad_actual=edad_actual,
        edad_jubilacion=edad_jubilacion,
        tasa_anual=perfil_riesgo,
    )
    ahorro_tradicional = None
    if ahorro_mensual_normal > 0:
        ahorro_tradicional = proyectar_jubilacion(
            ahorro_mensual=ahorro_mensual_normal,
            edad_actual=edad_actual,
            edad_jubilacion=edad_jubilacion,
            tasa_anual=perfil_riesgo,
            es_apv=False,
        )

    wb = Workbook()
    bold = Font(bold=True)

    # --- Hoja: Datos de entrada ---
    ws1 = wb.active
    ws1.title = "Datos"
    ws1.append(["Parámetro", "Valor"])
    ws1.append(["Sueldo bruto mensual", sueldo_bruto_mensual])
    ws1.append(["Edad actual", edad_actual])
    ws1.append(["Edad jubilación", edad_jubilacion])
    ws1.append(["Aporte APV mensual", ahorro_mensual_apv])
    ws1.append(["Ahorro normal mensual", ahorro_mensual_normal])
    ws1.append(["Tasa retorno anual (%)", round(perfil_riesgo * 100, 1)])
    for cell in ws1["A1:B1"][0]:
        cell.font = bold
    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 18

    # --- Hoja: Totales ---
    ws2 = wb.create_sheet("Totales")
    ws2.append(["Régimen", "Beneficio anual"])
    ws2["A1"].font = bold
    ws2["B1"].font = bold
    ra = resultado["regimen_a"]
    rb = resultado["regimen_b"]
    mix = resultado["mix"]
    ws2.append(["Régimen A (bonificación)", _fmt_num(ra["bonificacion_efectiva"])])
    ws2.append(["Régimen B (ahorro fiscal)", _fmt_num(rb["ahorro_fiscal"])])
    ws2.append(["Mix optimizado", _fmt_num(mix["beneficio_total"])])
    ws2.append([])
    ws2.append(["Mejor régimen", resultado["mejor_regimen"]])
    ws2["A5"].font = bold
    if ahorro_tradicional:
        ws2.append([])
        ws2.append(["Ahorro tradicional (capital neto a jubilación)", _fmt_num(ahorro_tradicional["capital_neto"])])
    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 18

    # --- Hoja: Proyección año a año ---
    ws3 = wb.create_sheet("Proyección")
    headers = ["Año", "Edad", "Capital APV", "Capital ahorro normal", "Aporte acum. APV", "Aporte acum. normal"]
    ws3.append(headers)
    for c in range(1, len(headers) + 1):
        ws3.cell(1, c).font = bold
    for row in proyeccion_anual:
        ws3.append([
            row["anio_proyeccion"],
            row["edad"],
            row["capital_apv"],
            row["capital_ahorro_tradicional"],
            row["aporte_acumulado_apv"],
            row["aporte_acumulado_normal"],
        ])
    for c in range(1, 7):
        ws3.column_dimensions[get_column_letter(c)].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
